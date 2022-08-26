import openpyxl
from openpyxl.utils import get_column_letter

import datetime
import pathlib

class SheetManager:
  def __init__(self, filename, sheetname) -> None:
    self.filename = filename
    self.sheetname = sheetname
    self.topRowOfData = 3
    self.lastColOfData = 11
    self.wb:openpyxl.Workbook = None
    self.ws = None

  def print_file_sheet_names(self):
    print(f"File Name: {self.filename}\nSheet Name: {self.sheetname}")

  def get_loads(self):
    self.wb = openpyxl.load_workbook(self.filename)
    self.ws = self.wb[self.sheetname]
    results = []
    for i, row in enumerate(self.ws.rows):
      row_loads = []
      if i < self.topRowOfData : continue
      for j, cell in enumerate(row):
        if j > self.lastColOfData: continue
        row_loads.append(cell.value)
      results.append(row_loads)
    return results

  def create_output_sheet(self,data = None):
    now = datetime.datetime.now()

    output_sheetname = f"{now.year}{now.month:0>2d}{now.day:0>2d}{now.hour:0>2d}{now.minute:0>2d}{now.second:0>2d}"
    filepath = pathlib.Path(self.filename)
    filename = filepath.with_name("_OUT_" + filepath.stem + filepath.suffix)
    output_filename = filename.resolve()
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = output_sheetname

    if data is None: return

    for i, row in enumerate(data,start=self.topRowOfData + 1):
      for j, val in enumerate(row, start=1):
        #_ = out_ws.cell(column=j, row=i, value="{0}".format( get_column_letter(j)))
        _ = out_ws.cell(column=j, row=i, value="{0}".format( val ))

    try:
      out_wb.save(output_filename)
    except Exception as e:
      raise e



def main():
  pass

if __name__ == "__main__":
  filename = r".\data\220825_section forces.xlsx"
  sheetname = r"Section Cut Forces - Analysis"
  sm = SheetManager(filename, sheetname)
  print(sm.get_loads())